import os
import json
import urllib.request
import urllib.parse
import subprocess
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tempfile

app = Flask(__name__)

# ==================== Config ====================
API_KEY = os.environ.get('YT_API_KEY', 'AIzaSyCzwAFKyLyGSr0jNWbTBVlE95rHe5z70H0')
PROXY_PORT = os.environ.get('YT_PROXY_PORT', '7897')
PROXY_URL = f'http://127.0.0.1:{PROXY_PORT}'
# On cloud (Render etc), no proxy needed. Set USE_PROXY=false to disable.
USE_PROXY = os.environ.get('USE_PROXY', 'true').lower() == 'true'
APP_PORT = int(os.environ.get('APP_PORT', os.environ.get('PORT', '5102')))

# TikTok API - SocialCrawl
TIKTOK_API_KEY = os.environ.get('TIKTOK_API_KEY', 'sc_UlUvDc4H9cF06kDtaqzMnEPJGicFM1WprfgJiLAzSlQ')
TIKTOK_API_BASE = 'https://www.socialcrawl.dev/v1/tiktok/search'

# TikHub API - 国内4平台 (抖音/小红书/B站/视频号)
TIKHUB_API_KEY = os.environ.get('TIKHUB_API_KEY', '')
TIKHUB_API_BASE = 'https://api.tikhub.dev'  # 国内免翻墙域名

YT_DLP_PATH = '/Users/zxl/.workbuddy/binaries/python/envs/default/bin/yt-dlp'


# ==================== Proxy Helpers ====================
def get_opener():
    if USE_PROXY:
        proxy_handler = urllib.request.ProxyHandler({
            'http': PROXY_URL,
            'https': PROXY_URL,
        })
        return urllib.request.build_opener(proxy_handler)
    else:
        return urllib.request.build_opener()


def api_request(url, timeout=20):
    opener = get_opener()
    req = urllib.request.Request(url)
    with opener.open(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode())


# ==================== Format Helpers ====================
def parse_duration(iso_duration):
    if not iso_duration or iso_duration == 'N/A':
        return 'N/A'
    match = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', iso_duration)
    if not match:
        # Try TikTok numeric seconds
        if iso_duration.isdigit():
            s = int(iso_duration)
            m, s = divmod(s, 60)
            h, m = divmod(m, 60)
            parts = []
            if h: parts.append(f'{h}小时')
            if m: parts.append(f'{m}分')
            if s: parts.append(f'{s}秒')
            return ''.join(parts) if parts else '0秒'
        return iso_duration
    h, m, s = match.groups()
    parts = []
    if h: parts.append(f'{h}小时')
    if m: parts.append(f'{m}分')
    if s: parts.append(f'{s}秒')
    return ''.join(parts) if parts else iso_duration


def format_number(num_str):
    if not num_str or num_str == 'N/A':
        return 'N/A'
    try:
        num = int(num_str)
        if num >= 10000:
            return f'{num/10000:.1f}万'
        return f'{num:,}'
    except:
        return num_str


def format_date(date_str):
    if not date_str or date_str == 'N/A':
        return 'N/A'
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except:
        # Try TikTok format YYYYMMDD
        if len(date_str) == 8 and date_str.isdigit():
            return f'{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}'
        return date_str


# ==================== YouTube Search ====================
def search_youtube(keyword, max_results=10, order='relevance', region=''):
    """Search YouTube and return unified video list."""
    try:
        search_url = (
            f'https://www.googleapis.com/youtube/v3/search'
            f'?part=snippet&type=video&maxResults={max_results}'
            f'&q={urllib.parse.quote(keyword)}&key={API_KEY}&order={order}'
        )
        if region:
            search_url += f'&regionCode={region}'

        data = api_request(search_url)
        items = data.get('items', [])

        if not items:
            return []

        video_ids = [item['id']['videoId'] for item in items]

        # Get stats
        stats_url = (
            f'https://www.googleapis.com/youtube/v3/videos'
            f'?part=statistics,contentDetails,snippet'
            f'&id={",".join(video_ids)}&key={API_KEY}'
        )
        stats_data = api_request(stats_url)

        # Get channel details
        channel_ids = list(set([item['snippet']['channelId'] for item in items]))
        channel_url = (
            f'https://www.googleapis.com/youtube/v3/channels'
            f'?part=statistics,snippet&id={",".join(channel_ids)}&key={API_KEY}'
        )
        channel_data = api_request(channel_url)

        channel_map = {}
        for ch in channel_data.get('items', []):
            channel_map[ch['id']] = {
                'subscriberCount': ch['statistics'].get('subscriberCount', 'N/A'),
                'videoCount': ch['statistics'].get('videoCount', 'N/A'),
                'totalViewCount': ch['statistics'].get('viewCount', 'N/A'),
                'channelAvatar': ch['snippet'].get('thumbnails', {}).get('default', {}).get('url', ''),
            }

        stats_map = {}
        for v in stats_data.get('items', []):
            tags = v.get('snippet', {}).get('tags', [])
            stats_map[v['id']] = {
                'viewCount': v['statistics'].get('viewCount', 'N/A'),
                'likeCount': v['statistics'].get('likeCount', 'N/A'),
                'commentCount': v['statistics'].get('commentCount', 'N/A'),
                'duration': v['contentDetails'].get('duration', 'N/A'),
                'tags': tags[:10] if tags else [],
            }

        videos = []
        for item in items:
            vid = item['id']['videoId']
            snippet = item['snippet']
            stats = stats_map.get(vid, {})
            channel_info = channel_map.get(snippet.get('channelId', ''), {})

            videos.append({
                'id': vid,
                'platform': 'youtube',
                'title': snippet.get('title', ''),
                'channelTitle': snippet.get('channelTitle', ''),
                'channelId': snippet.get('channelId', ''),
                'channelAvatar': channel_info.get('channelAvatar', ''),
                'subscriberCount': channel_info.get('subscriberCount', 'N/A'),
                'channelVideoCount': channel_info.get('videoCount', 'N/A'),
                'channelTotalViews': channel_info.get('totalViewCount', 'N/A'),
                'publishedAt': snippet.get('publishedAt', ''),
                'publishedAtFormatted': format_date(snippet.get('publishedAt', '')),
                'viewCount': stats.get('viewCount', 'N/A'),
                'viewCountFormatted': format_number(stats.get('viewCount', 'N/A')),
                'likeCount': stats.get('likeCount', 'N/A'),
                'likeCountFormatted': format_number(stats.get('likeCount', 'N/A')),
                'commentCount': stats.get('commentCount', 'N/A'),
                'commentCountFormatted': format_number(stats.get('commentCount', 'N/A')),
                'duration': stats.get('duration', ''),
                'durationFormatted': parse_duration(stats.get('duration', '')),
                'description': snippet.get('description', ''),
                'thumbnail': snippet.get('thumbnails', {}).get('high', {}).get('url', ''),
                'tags': stats.get('tags', []),
                'url': f'https://youtube.com/watch?v={vid}',
            })
        return videos

    except Exception as e:
        print(f'YouTube search error: {e}')
        return []


# ==================== TikTok Search (SocialCrawl API) ====================
def search_tiktok(keyword, max_results=10):
    """Search TikTok via SocialCrawl API.
    
    SocialCrawl provides keyword-based TikTok video search.
    Free tier: 400 credits, no credit card required.
    API docs: https://www.socialcrawl.dev/platforms/tiktok/search
    """
    if not TIKTOK_API_KEY:
        print('TikTok API Key not configured')
        return []

    try:
        params = {
            'query': keyword,
            'sort_by': 'relevance',
        }
        url = f'{TIKTOK_API_BASE}?{urllib.parse.urlencode(params)}'

        req = urllib.request.Request(url)
        req.add_header('x-api-key', TIKTOK_API_KEY)
        req.add_header('Accept', 'application/json')

        opener = get_opener()
        with opener.open(req, timeout=30) as resp:
            data = json.loads(resp.read().decode('utf-8'))

        items = data.get('data', {}).get('items', [])

        videos = []
        for item in items[:max_results]:
            aweme = item.get('aweme_info', {})

            # Extract key fields
            video_id = aweme.get('aweme_id', '')
            desc = aweme.get('desc', '')

            # Author
            author = aweme.get('author', {})
            author_name = author.get('nickname', '')
            author_id = author.get('unique_id', author.get('uid', ''))

            # Stats
            stats = aweme.get('statistics', {})
            plays = stats.get('play_count', 0)
            likes = stats.get('digg_count', 0)
            comments = stats.get('comment_count', 0)
            shares = stats.get('share_count', 0)
            collects = stats.get('collect_count', 0)

            # Thumbnail
            cover = aweme.get('video', {}).get('cover', {})
            thumb_urls = cover.get('url_list', [])
            thumbnail = thumb_urls[-1] if thumb_urls else ''

            # Duration (milliseconds)
            duration_ms = aweme.get('video', {}).get('duration', 0)
            duration_sec = round(duration_ms / 1000) if duration_ms else 0

            # Create time (unix timestamp)
            create_time = aweme.get('create_time', 0)
            if create_time:
                pub_date = datetime.utcfromtimestamp(create_time)
                pub_date_str = pub_date.strftime('%Y-%m-%dT%H:%M:%SZ')
                pub_date_formatted = pub_date.strftime('%Y-%m-%d')
            else:
                pub_date_str = ''
                pub_date_formatted = ''

            # Hashtags
            text_extra = aweme.get('text_extra', [])
            hashtags = [t.get('hashtag_name', '') for t in text_extra if t.get('hashtag_name')]

            videos.append({
                'id': video_id,
                'platform': 'tiktok',
                'title': desc[:200] if desc else '',
                'channelTitle': f'{author_name} (@{author_id})' if author_id else author_name,
                'channelId': author_id,
                'channelAvatar': '',
                'subscriberCount': 'N/A',
                'channelVideoCount': 'N/A',
                'channelTotalViews': 'N/A',
                'publishedAt': pub_date_str,
                'publishedAtFormatted': pub_date_formatted,
                'viewCount': str(plays) if plays else '0',
                'viewCountFormatted': format_number(str(plays)),
                'likeCount': str(likes) if likes else '0',
                'likeCountFormatted': format_number(str(likes)),
                'commentCount': str(comments) if comments else '0',
                'commentCountFormatted': format_number(str(comments)),
                'shareCount': str(shares) if shares else '0',
                'shareCountFormatted': format_number(str(shares)),
                'collectCount': str(collects) if collects else '0',
                'duration': str(duration_sec),
                'durationFormatted': parse_duration(str(duration_sec)),
                'description': desc,
                'thumbnail': thumbnail,
                'tags': hashtags[:10],
                'url': f'https://www.tiktok.com/@{author_id}/video/{video_id}' if author_id and video_id else '',
            })

        return videos

    except urllib.error.HTTPError as e:
        error_body = ''
        try:
            error_body = e.read().decode('utf-8', errors='replace')[:200]
        except:
            pass
        print(f'TikTok search HTTP error: {e.code} {e.reason} - {error_body}')
        return []
    except Exception as e:
        print(f'TikTok search error: {e}')
        return []


# ==================== TikHub API Helpers ====================
def tikhub_get(path, api_key, params=None):
    """Call TikHub GET endpoint (no proxy needed for api.tikhub.dev)"""
    url = f'{TIKHUB_API_BASE}{path}'
    if params:
        url += '?' + urllib.parse.urlencode(params)

    req = urllib.request.Request(url)
    req.add_header('Authorization', f'Bearer {api_key}')
    req.add_header('Accept', 'application/json')

    import ssl
    ctx = ssl.create_default_context()
    opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=ctx))

    with opener.open(req, timeout=30) as resp:
        result = json.loads(resp.read().decode('utf-8'))
        # data 字段可能是 JSON 字符串，需二次解析
        if result.get('data') and isinstance(result['data'], str):
            try:
                result['data'] = json.loads(result['data'])
            except:
                pass
        return result


def tikhub_post(path, api_key, body=None):
    """Call TikHub POST endpoint (no proxy needed for api.tikhub.dev)"""
    url = f'{TIKHUB_API_BASE}{path}'
    data = json.dumps(body or {}).encode('utf-8')

    req = urllib.request.Request(url, data=data, method='POST')
    req.add_header('Authorization', f'Bearer {api_key}')
    req.add_header('Content-Type', 'application/json')
    req.add_header('Accept', 'application/json')

    import ssl
    ctx = ssl.create_default_context()
    opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=ctx))

    with opener.open(req, timeout=30) as resp:
        result = json.loads(resp.read().decode('utf-8'))
        if result.get('data') and isinstance(result['data'], str):
            try:
                result['data'] = json.loads(result['data'])
            except:
                pass
        return result


# ==================== Douyin Search (TikHub) ====================
def search_douyin(keyword, max_results=10):
    """Search Douyin (抖音) videos via TikHub API."""
    if not TIKHUB_API_KEY:
        print('TikHub API Key not configured, skip Douyin')
        return []

    try:
        result = tikhub_post(
            '/api/v1/douyin/search/fetch_video_search_v1',
            TIKHUB_API_KEY,
            body={
                'keyword': keyword,
                'cursor': 0,
                'sort_type': '0',
                'publish_time': '0',
                'filter_duration': '0',
                'content_type': '0',
                'search_id': '',
                'backtrace': '',
            }
        )

        data = result.get('data', {})
        items = data.get('data', []) if isinstance(data, dict) else []

        videos = []
        for item in items[:max_results]:
            aweme = item.get('aweme_info', {})
            if not aweme:
                continue

            author = aweme.get('author', {})
            stats = aweme.get('statistics', {})
            cover_urls = aweme.get('video', {}).get('cover', {}).get('url_list', [])
            duration_ms = aweme.get('video', {}).get('duration', 0)
            duration_sec = round(duration_ms / 1000) if duration_ms else 0
            create_time = aweme.get('create_time', 0)

            if create_time:
                pub_date = datetime.utcfromtimestamp(create_time)
                pub_date_str = pub_date.strftime('%Y-%m-%dT%H:%M:%SZ')
                pub_date_formatted = pub_date.strftime('%Y-%m-%d')
            else:
                pub_date_str = ''
                pub_date_formatted = ''

            text_extra = aweme.get('text_extra', [])
            hashtags = [t.get('hashtag_name', '') for t in text_extra if t.get('hashtag_name')]

            videos.append({
                'id': aweme.get('aweme_id', ''),
                'platform': 'douyin',
                'title': (aweme.get('desc', '') or '')[:200],
                'channelTitle': author.get('nickname', ''),
                'channelId': author.get('unique_id', author.get('uid', '')),
                'channelAvatar': '',
                'subscriberCount': str(author.get('follower_count', 'N/A')),
                'channelVideoCount': 'N/A',
                'channelTotalViews': 'N/A',
                'publishedAt': pub_date_str,
                'publishedAtFormatted': pub_date_formatted,
                'viewCount': str(stats.get('play_count', 0)),
                'viewCountFormatted': format_number(str(stats.get('play_count', 0))),
                'likeCount': str(stats.get('digg_count', 0)),
                'likeCountFormatted': format_number(str(stats.get('digg_count', 0))),
                'commentCount': str(stats.get('comment_count', 0)),
                'commentCountFormatted': format_number(str(stats.get('comment_count', 0))),
                'shareCount': str(stats.get('share_count', 0)),
                'shareCountFormatted': format_number(str(stats.get('share_count', 0))),
                'collectCount': str(stats.get('collect_count', 0)),
                'duration': str(duration_sec),
                'durationFormatted': parse_duration(str(duration_sec)),
                'description': aweme.get('desc', ''),
                'thumbnail': cover_urls[-1] if cover_urls else '',
                'tags': hashtags[:10],
                'url': aweme.get('share_url', ''),
            })

        return videos

    except urllib.error.HTTPError as e:
        print(f'Douyin search HTTP error: {e.code} {e.reason}')
        return []
    except Exception as e:
        print(f'Douyin search error: {e}')
        return []


# ==================== Xiaohongshu Search (TikHub) ====================
def search_xiaohongshu(keyword, max_results=10):
    """Search Xiaohongshu (小红书) notes via TikHub API."""
    if not TIKHUB_API_KEY:
        print('TikHub API Key not configured, skip Xiaohongshu')
        return []

    try:
        result = tikhub_get(
            '/api/v1/xiaohongshu/web/search_notes',
            TIKHUB_API_KEY,
            params={
                'keyword': keyword,
                'page': 1,
                'sort': 'general',
                'noteType': '_0',
            }
        )

        data = result.get('data', {})
        items = data.get('items', []) if isinstance(data, dict) else []

        videos = []
        for item in items[:max_results]:
            note = item.get('note_card', item) if isinstance(item, dict) else {}
            if not isinstance(note, dict):
                continue

            user = note.get('user', {})
            interact = note.get('interact_info', {})
            cover = note.get('cover', {})

            # Thumbnail
            thumb_url = ''
            if isinstance(cover, dict):
                url_list = cover.get('url_list', cover.get('url', ''))
                if isinstance(url_list, list) and url_list:
                    thumb_url = url_list[-1]
                elif isinstance(url_list, str):
                    thumb_url = url_list

            # Tags
            tag_list = note.get('tag_list', [])
            tags = [t.get('name', '') for t in tag_list if isinstance(t, dict) and t.get('name')]

            # Note type
            note_type = note.get('type', note.get('note_type', ''))
            type_label = '视频' if str(note_type) in ('2', 'video') else '图文'

            # Duration
            video_info = note.get('video', {})
            duration_str = ''
            if isinstance(video_info, dict):
                duration_val = video_info.get('duration', 0)
                if duration_val:
                    duration_str = str(round(float(duration_val)))

            videos.append({
                'id': note.get('note_id', note.get('id', '')),
                'platform': 'xiaohongshu',
                'title': (note.get('title', note.get('display_title', '')) or '')[:200],
                'channelTitle': user.get('nickname', ''),
                'channelId': user.get('user_id', user.get('userid', '')),
                'channelAvatar': '',
                'subscriberCount': 'N/A',
                'channelVideoCount': 'N/A',
                'channelTotalViews': 'N/A',
                'publishedAt': note.get('time', ''),
                'publishedAtFormatted': note.get('time', ''),
                'viewCount': '0',
                'viewCountFormatted': 'N/A',
                'likeCount': interact.get('liked_count', '0'),
                'likeCountFormatted': interact.get('liked_count', '0'),
                'commentCount': interact.get('comment_count', '0'),
                'commentCountFormatted': interact.get('comment_count', '0'),
                'shareCount': interact.get('share_count', '0'),
                'shareCountFormatted': interact.get('share_count', '0'),
                'collectCount': interact.get('collected_count', '0'),
                'duration': duration_str,
                'durationFormatted': parse_duration(duration_str) if duration_str else type_label,
                'description': note.get('desc', ''),
                'thumbnail': thumb_url,
                'tags': tags[:10],
                'url': f"https://www.xiaohongshu.com/explore/{note.get('note_id', '')}",
            })

        return videos

    except urllib.error.HTTPError as e:
        print(f'Xiaohongshu search HTTP error: {e.code} {e.reason}')
        return []
    except Exception as e:
        print(f'Xiaohongshu search error: {e}')
        return []


# ==================== Bilibili Search (TikHub) ====================
def search_bilibili(keyword, max_results=10):
    """Search Bilibili (B站) videos via TikHub API."""
    if not TIKHUB_API_KEY:
        print('TikHub API Key not configured, skip Bilibili')
        return []

    try:
        result = tikhub_get(
            '/api/v1/bilibili/web/fetch_general_search',
            TIKHUB_API_KEY,
            params={
                'keyword': keyword,
                'order': 'totalrank',
                'page': 1,
                'page_size': max_results,
            }
        )

        data = result.get('data', {})
        # B站搜索结果可能在 data.result 中
        items = []
        if isinstance(data, dict):
            result_data = data.get('result', [])
            if isinstance(result_data, list):
                items = result_data
            elif isinstance(result_data, dict):
                items = result_data.get('result', result_data.get('video', []))
                if not isinstance(items, list):
                    items = []

        videos = []
        for item in items[:max_results]:
            if not isinstance(item, dict):
                continue

            # Clean HTML tags from title
            title = re.sub(r'<[^>]+>', '', item.get('title', ''))

            # Stats
            play = item.get('play', 0)
            if isinstance(play, str) and not play.isdigit():
                play = 0
            like = item.get('like', 0)
            danmaku = item.get('video_review', item.get('danmaku', 0))

            # Duration
            duration_str = item.get('duration', item.get('length', ''))
            duration_sec = 0
            if duration_str and ':' in str(duration_str):
                parts = str(duration_str).split(':')
                try:
                    if len(parts) == 2:
                        duration_sec = int(parts[0]) * 60 + int(parts[1])
                    elif len(parts) == 3:
                        duration_sec = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                except:
                    pass

            # Tags
            tag_str = item.get('tag', '')
            tags = [t.strip() for t in tag_str.split(',')] if tag_str else []

            # BVID
            bvid = item.get('bvid', '')
            aid = item.get('aid', '')

            videos.append({
                'id': bvid or str(aid),
                'platform': 'bilibili',
                'title': title[:200],
                'channelTitle': item.get('author', ''),
                'channelId': item.get('mid', ''),
                'channelAvatar': '',
                'subscriberCount': 'N/A',
                'channelVideoCount': 'N/A',
                'channelTotalViews': 'N/A',
                'publishedAt': str(item.get('pubdate', '')),
                'publishedAtFormatted': str(item.get('pubdate', '')),
                'viewCount': str(play),
                'viewCountFormatted': format_number(str(play)),
                'likeCount': str(like),
                'likeCountFormatted': format_number(str(like)),
                'commentCount': str(danmaku),
                'commentCountFormatted': format_number(str(danmaku)),
                'shareCount': '0',
                'shareCountFormatted': 'N/A',
                'collectCount': str(item.get('favorites', 0)),
                'duration': str(duration_sec),
                'durationFormatted': str(duration_str) if duration_str else 'N/A',
                'description': item.get('description', ''),
                'thumbnail': item.get('pic', ''),
                'tags': tags[:10],
                'url': f'https://www.bilibili.com/video/{bvid}' if bvid else f'https://www.bilibili.com/video/av{aid}',
            })

        return videos

    except urllib.error.HTTPError as e:
        print(f'Bilibili search HTTP error: {e.code} {e.reason}')
        return []
    except Exception as e:
        print(f'Bilibili search error: {e}')
        return []


# ==================== WeChat Channels Search (TikHub) ====================
def search_wechat_channels(keyword, max_results=10):
    """Search WeChat Channels (视频号) via TikHub API."""
    if not TIKHUB_API_KEY:
        print('TikHub API Key not configured, skip WeChat Channels')
        return []

    try:
        result = tikhub_get(
            '/api/v1/wechat_channels/fetch_search_channels',
            TIKHUB_API_KEY,
            params={
                'keyword': keyword,
                'offset': 0,
                'sort_type': '_0',
            }
        )

        data = result.get('data', {})
        items = data.get('object_list', []) if isinstance(data, dict) else []

        videos = []
        for item in items[:max_results]:
            if not isinstance(item, dict):
                continue

            object_desc = item.get('object_desc', {})
            finder_user = item.get('finder_user', {})
            social_info = item.get('social_info', {})
            media = object_desc.get('media', [])

            # Thumbnail
            thumb_url = ''
            if media and isinstance(media, list):
                thumb_url = media[0].get('cover_url', media[0].get('url', ''))

            # Duration
            duration_val = object_desc.get('duration', 0)
            duration_sec = 0
            try:
                duration_sec = int(float(duration_val))
            except:
                pass

            videos.append({
                'id': object_desc.get('object_id', ''),
                'platform': 'wechat',
                'title': (object_desc.get('description', '') or '')[:200],
                'channelTitle': finder_user.get('nickname', ''),
                'channelId': finder_user.get('username', ''),
                'channelAvatar': finder_user.get('head_url', ''),
                'subscriberCount': 'N/A',
                'channelVideoCount': 'N/A',
                'channelTotalViews': 'N/A',
                'publishedAt': str(object_desc.get('create_time', '')),
                'publishedAtFormatted': str(object_desc.get('create_time', '')),
                'viewCount': str(social_info.get('play_count', social_info.get('read_count', 0))),
                'viewCountFormatted': format_number(str(social_info.get('play_count', social_info.get('read_count', 0)))),
                'likeCount': str(social_info.get('like_count', social_info.get('liked_count', 0))),
                'likeCountFormatted': format_number(str(social_info.get('like_count', social_info.get('liked_count', 0)))),
                'commentCount': str(social_info.get('comment_count', 0)),
                'commentCountFormatted': format_number(str(social_info.get('comment_count', 0))),
                'shareCount': str(social_info.get('share_count', 0)),
                'shareCountFormatted': format_number(str(social_info.get('share_count', 0))),
                'collectCount': '0',
                'duration': str(duration_sec),
                'durationFormatted': f'{duration_sec}s' if duration_sec else 'N/A',
                'description': object_desc.get('description', ''),
                'thumbnail': thumb_url,
                'tags': [],
                'url': '',
            })

        return videos

    except urllib.error.HTTPError as e:
        print(f'WeChat Channels search HTTP error: {e.code} {e.reason}')
        return []
    except Exception as e:
        print(f'WeChat Channels search error: {e}')
        return []


# ==================== Routes ====================
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response

@app.route('/<path:path>', methods=['OPTIONS'])
@app.route('/', methods=['OPTIONS'])
def options_handler(path=''):
    return '', 204

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/search')
def search():
    """Search single platform (backward compatible) or multi-platform."""
    keyword = request.args.get('q', '').strip()
    max_results = min(int(request.args.get('max', 10)), 50)
    order = request.args.get('order', 'relevance')
    region = request.args.get('region', '')
    tikhub_key = request.args.get('tikhub_key', '').strip()
    platforms = request.args.get('platforms', 'youtube,tiktok').split(',')

    if not keyword:
        return jsonify({'error': '请输入搜索关键词'}), 400

    # Override TikHub key if provided in request
    global TIKHUB_API_KEY
    saved_tikhub_key = TIKHUB_API_KEY
    if tikhub_key:
        TIKHUB_API_KEY = tikhub_key

    all_videos = []
    errors = {}

    # Platform search functions mapping
    PLATFORM_FUNCS = {
        'youtube': lambda: search_youtube(keyword, max_results, order, region),
        'tiktok': lambda: search_tiktok(keyword, max_results),
        'douyin': lambda: search_douyin(keyword, max_results),
        'xiaohongshu': lambda: search_xiaohongshu(keyword, max_results),
        'bilibili': lambda: search_bilibili(keyword, max_results),
        'wechat': lambda: search_wechat_channels(keyword, max_results),
    }

    # Parallel search
    active_platforms = [p for p in platforms if p in PLATFORM_FUNCS]
    with ThreadPoolExecutor(max_workers=min(len(active_platforms), 6)) as executor:
        futures = {}
        for platform in active_platforms:
            futures[executor.submit(PLATFORM_FUNCS[platform])] = platform

        for future in as_completed(futures):
            platform = futures[future]
            try:
                result = future.result()
                all_videos.extend(result)
            except Exception as e:
                errors[platform] = str(e)

    # Restore original key
    TIKHUB_API_KEY = saved_tikhub_key

    # Sort by view count desc if order is viewCount
    if order == 'viewCount':
        all_videos.sort(key=lambda x: int(x.get('viewCount', '0') or '0'), reverse=True)

    return jsonify({
        'videos': all_videos,
        'total': len(all_videos),
        'keyword': keyword,
        'errors': errors if errors else None,
    })


@app.route('/api/export', methods=['POST'])
def export_excel():
    """Export raw video search results to Excel."""
    data = request.json
    videos = data.get('videos', [])
    keyword = data.get('keyword', '搜索结果')

    if not videos:
        return jsonify({'error': '没有数据可导出'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '视频数据'

    # Headers
    headers = ['序号', '平台', '标题', '频道名', '订阅数', '发布日期', '播放量', '点赞数', '评论数', '时长', '标签', '视频链接']
    header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='d0d0d0'),
        right=Side(style='thin', color='d0d0d0'),
        top=Side(style='thin', color='d0d0d0'),
        bottom=Side(style='thin', color='d0d0d0'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    for i, v in enumerate(videos, 1):
        platform_name = {'youtube': 'YouTube', 'tiktok': 'TikTok', 'douyin': '抖音',
                         'xiaohongshu': '小红书', 'bilibili': 'B站', 'wechat': '视频号'}.get(v.get('platform'), v.get('platform', ''))
        row_data = [
            i,
            platform_name,
            v.get('title', ''),
            v.get('channelTitle', ''),
            v.get('subscriberCount', 'N/A'),
            v.get('publishedAtFormatted', ''),
            v.get('viewCount', 'N/A'),
            v.get('likeCount', 'N/A'),
            v.get('commentCount', 'N/A'),
            v.get('durationFormatted', ''),
            ', '.join(v.get('tags', [])),
            v.get('url', ''),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.font = Font(name='微软雅黑', size=10)
            cell.border = thin_border
            if col == 3:  # title
                cell.alignment = Alignment(wrap_text=True)
            if col == 12:  # url
                cell.font = Font(name='微软雅黑', size=10, color='0563C1', underline='single')

    # Column widths
    widths = [6, 10, 50, 20, 14, 14, 12, 12, 12, 12, 30, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:L{len(videos)+1}'

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f'视频搜索_{keyword}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== AI Analysis ====================
@app.route('/api/analyze', methods=['POST'])
def analyze_topics():
    """AI analyze videos and generate topic library."""
    data = request.json
    videos = data.get('videos', [])
    keyword = data.get('keyword', '')
    llm_api_key = data.get('llm_api_key', '')
    llm_model = data.get('llm_model', 'gpt-4o-mini')
    llm_base_url = data.get('llm_base_url', 'https://api.openai.com/v1')

    if not videos:
        return jsonify({'error': '没有视频数据可分析'}), 400

    if not llm_api_key:
        return jsonify({'error': '请配置 LLM API Key'}), 400

    try:
        # Build analysis prompt
        video_summaries = []
        for i, v in enumerate(videos[:30], 1):  # Limit to 30 videos
            platform_map = {
                'youtube': 'YouTube', 'tiktok': 'TikTok', 'douyin': '抖音',
                'xiaohongshu': '小红书', 'bilibili': 'B站', 'wechat': '视频号'
            }
            platform = platform_map.get(v.get('platform', ''), v.get('platform', ''))
            video_summaries.append(
                f"[{i}] 平台: {platform}\n"
                f"    标题: {v.get('title', '')}\n"
                f"    作者: {v.get('channelTitle', '')}\n"
                f"    播放量: {v.get('viewCountFormatted', 'N/A')}\n"
                f"    点赞: {v.get('likeCountFormatted', 'N/A')}\n"
                f"    标签: {', '.join(v.get('tags', [])[:5])}\n"
                f"    简介: {v.get('description', '')[:200]}\n"
            )

        prompt = f"""你是一位资深的内容营销专家和跨境电商自媒体顾问。

请基于以下从 YouTube、TikTok、抖音、小红书、B站、视频号 搜索到的关于"{keyword}"的视频数据，进行深度分析，生成一份结构化选题库。

## 分析要求
1. **话题聚类**：将视频按主题聚类，识别出 3-8 个核心热门话题
2. **梗概总结**：每个话题总结其为什么火（痛点/需求/趋势），核心内容是什么
3. **仿写建议**：每个话题给出 3-5 个可以直接使用的仿写标题建议
4. **内容框架**：每个话题给出内容创作的框架建议（开头怎么抓眼球、中间怎么展开、结尾怎么引导）
5. **热度评级**：基于播放量数据，给出高/中/低的热度评级
6. **最优平台**：分析该话题在哪个平台（YouTube/TikTok）表现更好

## 视频数据
{chr(10).join(video_summaries)}

## 输出格式
请严格按以下 JSON 格式输出，不要添加任何 markdown 标记或额外说明：

{{
  "topics": [
    {{
      "topic_name": "话题名称",
      "summary": "话题梗概：为什么这个话题火，核心内容是什么",
      "heat_level": "高|中|低",
      "best_platform": "YouTube|TikTok|抖音|小红书|B站|视频号|多平台",
      "rewrite_titles": ["仿写标题1", "仿写标题2", "仿写标题3", "仿写标题4", "仿写标题5"],
      "content_framework": {{
        "hook": "开头怎么吸引注意力（30字以内）",
        "body": "中间内容展开思路（50字以内）",
        "cta": "结尾引导/行动号召（30字以内）"
      }},
      "key_insights": ["洞察1", "洞察2", "洞察3"],
      "example_videos": [
        {{"title": "示例视频标题", "platform": "YouTube|TikTok", "url": "视频链接", "views": "播放量"}}
      ]
    }}
  ]
}}
"""

        # Call LLM API
        req_data = {
            'model': llm_model,
            'messages': [
                {'role': 'system', 'content': 'You are an expert content marketing analyst. Always respond with valid JSON only.'},
                {'role': 'user', 'content': prompt}
            ],
            'temperature': 0.7,
            'max_tokens': 4000,
        }

        req = urllib.request.Request(
            f'{llm_base_url.rstrip("/")}/chat/completions',
            data=json.dumps(req_data).encode('utf-8'),
            headers={
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {llm_api_key}',
            },
            method='POST'
        )

        opener = get_opener()
        try:
            with opener.open(req, timeout=90) as resp:
                llm_response = json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8', errors='replace') if e.fp else ''
            try:
                err_json = json.loads(error_body)
                err_msg = err_json.get('error', {}).get('message', error_body[:200])
            except:
                err_msg = error_body[:200] if error_body else str(e)
            return jsonify({'error': f'LLM API 错误 ({e.code}): {err_msg}'}), 502
        except urllib.error.URLError as e:
            return jsonify({'error': f'无法连接 LLM API，请检查 API 地址和代理设置: {str(e.reason)}'}), 502

        content = llm_response['choices'][0]['message']['content']

        # Extract JSON from response (in case there's markdown)
        json_match = re.search(r'\{[\s\S]*\}', content)
        if json_match:
            content = json_match.group()

        result = json.loads(content)
        topics = result.get('topics', [])

        # Enrich example videos with actual data
        platform_map = {
            'youtube': 'YouTube', 'tiktok': 'TikTok', 'douyin': '抖音',
            'xiaohongshu': '小红书', 'bilibili': 'B站', 'wechat': '视频号'
        }
        for topic in topics:
            for ex in topic.get('example_videos', []):
                # Find matching video
                for v in videos:
                    if v.get('title') == ex.get('title') or v.get('url') == ex.get('url'):
                        ex['url'] = v.get('url', '')
                        ex['platform'] = platform_map.get(v.get('platform', ''), v.get('platform', ''))
                        ex['views'] = v.get('viewCountFormatted', 'N/A')
                        break

        return jsonify({
            'topics': topics,
            'keyword': keyword,
            'total_videos_analyzed': len(videos),
        })

    except json.JSONDecodeError as e:
        return jsonify({'error': f'AI 返回格式解析失败: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'AI 分析失败: {str(e)}'}), 500


@app.route('/api/export_topics', methods=['POST'])
def export_topics_excel():
    """Export topic library to Excel."""
    data = request.json
    topics = data.get('topics', [])
    keyword = data.get('keyword', '选题库')

    if not topics:
        return jsonify({'error': '没有选题数据可导出'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '选题库'

    # Headers
    headers = ['序号', '话题名称', '热度', '最优平台', '话题梗概', '仿写标题1', '仿写标题2', '仿写标题3', '仿写标题4', '仿写标题5', '开头钩子', '内容展开', '结尾引导', '关键洞察']
    header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='d0d0d0'),
        right=Side(style='thin', color='d0d0d0'),
        top=Side(style='thin', color='d0d0d0'),
        bottom=Side(style='thin', color='d0d0d0'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for i, topic in enumerate(topics, 1):
        framework = topic.get('content_framework', {})
        titles = topic.get('rewrite_titles', [])
        insights = topic.get('key_insights', [])

        row_data = [
            i,
            topic.get('topic_name', ''),
            topic.get('heat_level', ''),
            topic.get('best_platform', ''),
            topic.get('summary', ''),
            titles[0] if len(titles) > 0 else '',
            titles[1] if len(titles) > 1 else '',
            titles[2] if len(titles) > 2 else '',
            titles[3] if len(titles) > 3 else '',
            titles[4] if len(titles) > 4 else '',
            framework.get('hook', ''),
            framework.get('body', ''),
            framework.get('cta', ''),
            ' | '.join(insights),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.font = Font(name='微软雅黑', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    widths = [6, 25, 8, 12, 50, 30, 30, 30, 30, 30, 25, 35, 25, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:N{len(topics)+1}'

    # Add example videos sheet
    if any(t.get('example_videos') for t in topics):
        ws2 = wb.create_sheet('关联视频')
        headers2 = ['话题', '平台', '视频标题', '播放量', '链接']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        row_idx = 2
        for topic in topics:
            for ex in topic.get('example_videos', []):
                ws2.cell(row=row_idx, column=1, value=topic.get('topic_name', '')).border = thin_border
                ws2.cell(row=row_idx, column=2, value=ex.get('platform', '')).border = thin_border
                ws2.cell(row=row_idx, column=3, value=ex.get('title', '')).border = thin_border
                ws2.cell(row=row_idx, column=4, value=ex.get('views', '')).border = thin_border
                link_cell = ws2.cell(row=row_idx, column=5, value=ex.get('url', ''))
                link_cell.border = thin_border
                link_cell.font = Font(name='微软雅黑', size=10, color='0563C1', underline='single')
                row_idx += 1

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 45
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 45

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f'选题库_{keyword}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== Main ====================
if __name__ == '__main__':
    print(f"🚀 多平台选题搜索工具启动 (YouTube/TikTok/抖音/小红书/B站/视频号)")
    print(f"📡 代理端口: {PROXY_PORT}")
    print(f"🔑 YouTube API Key: {API_KEY[:10]}...")
    print(f"🔑 TikTok API Key: {TIKTOK_API_KEY[:10]}...")
    print(f"🔑 TikHub API Key: {'已配置' if TIKHUB_API_KEY else '未配置(国内平台不可用)'}")
    print(f"🌐 访问地址: http://localhost:{APP_PORT}")
    app.run(host='0.0.0.0', port=APP_PORT, debug=False)
